import openpyxl
import os
import pickle

class make_list_of_links_per_domain():

    def __init__(self):
        self.list_of_agriculture_links = []
        self.list_of_finance_links = []
        self.list_of_general_links = []

    def make_list_of_domain_specific_links(self, excel_filename):
        """
        Parse excel sheet and collect links into lists categorized acc to domain
        """

        # Collecting all links from the excel sheet and adding them to lists acc to domain
        my_workbook = openpyxl.load_workbook(excel_filename)
        for sheet_num, sheet in enumerate(my_workbook.sheetnames):
            worksheet = my_workbook[sheet]
            if(sheet_num == 0):
                for row in worksheet.iter_rows(min_col=5):
                    for cell in row:
                        try:
                            link = cell.hyperlink.target
                        except:
                            link = None
                        if(link is not None):
                            self.list_of_agriculture_links.append(link)
                with open('agriculture_list_of_links', 'wb') as fp:
                    pickle.dump(self.list_of_agriculture_links, fp)
                print("Agriculture: done!")

            if(sheet_num == 1):
                for row in worksheet.iter_rows(min_col=5):
                    for cell in row:
                        try:
                            link = cell.hyperlink.target
                        except:
                            link = None
                        if(link is not None):
                            self.list_of_finance_links.append(link)
                with open('finance_list_of_links', 'wb') as fp:
                    pickle.dump(self.list_of_finance_links, fp)                
                print("finance: done!")

            if(sheet_num == 2):
                for row in worksheet.iter_rows(min_col=5):
                    for cell in row:
                        try:
                            link = cell.hyperlink.target
                        except:
                            link = None
                        if(link is not None):
                            self.list_of_general_links.append(link)
                with open('general_list_of_links', 'wb') as fp:
                    pickle.dump(self.list_of_general_links, fp)
                print("general: done!")

        # print(f"The number of links collected is {len(self.list_of_links)}.")
        # self.list_of_links.sort(key= self.list_of_links_sorting_utility)
link_organizer = make_list_of_links_per_domain()
link_organizer.make_list_of_domain_specific_links(excel_filename= 'HINDI-HEMANTHA.xlsx')