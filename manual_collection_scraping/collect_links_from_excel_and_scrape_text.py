import pickle
import openpyxl
import webbrowser
from importlib import reload
import scrape_text
import os

class DataScraperFromLinks():

    def __init__(self):
        self.list_of_links = []
        self.url_dict = {}

    def list_of_links_sorting_utility(self, link):
        """
        Function to remove the scheme part of the URL(eg: http, https etc) so that sorting can take place correctly.
        """
        partitioned_link_tuple = link.partition('/')
        link_without_scheme = partitioned_link_tuple[2]
        return link_without_scheme

    def copy_links_from_excel_and_scrape_text(self, excel_filename):
        """
        This function collects links from an excel spreadsheet and appends them to a Python list. Then it groups links from same websites together, so that you can go through each website, write the scraper function for it, and then collect text from all links of the same website together.
        """
        # Collecting all links from the excel sheet and sorting them to group same website's links together
        my_workbook = openpyxl.load_workbook(excel_filename)
        for sheet in my_workbook.sheetnames:
            worksheet = my_workbook[sheet]
            for row in worksheet.iter_rows(min_col=5):
                for cell in row:
                    try:
                        link = cell.hyperlink.target
                    except:
                        link = None
                    if(link is not None):
                        self.list_of_links.append(link)
        print(f"The number of links collected is {len(self.list_of_links)}.")

        ## temporary  code
        # with open('finance_list_of_links', 'rb') as fp:
        #     self.list_of_links = pickle.load(fp)

        self.list_of_links.sort(key= self.list_of_links_sorting_utility)
        
        # Making a dictionary of websites with key value pairs such that the name of website is key and the number of links from that website is value

        prev_link_website = None
        for each_link in self.list_of_links:
            split_link = each_link.split("/")
            website = split_link[2]
            # print(f"current website: {website}")
            # print(f"prev website: {prev_link_website}")
            if(website != prev_link_website):
                count = 1
                self.url_dict.update({website : count})
            if(website == prev_link_website):
                count += 1
                self.url_dict.update({website : count})
            # if(website == "apeda.gov.in"):
            #     print(each_link)
            prev_link_website = website
        print(f"The number of websites are: {len(self.url_dict)}")
        link_exceptions = []
        self.url_dict = {k: v for k, v in sorted(self.url_dict.items(), key=lambda item: item[1])}
        print(self.url_dict)
        # print(self.list_of_links)
        # for link in self.list_of_links:
        #     print(link[:30])


        
        # Iterate through the links acc to dictionary and for each website define the text scraping and scrape text
        j = 0    
        k = 2017 #change k to value in "position_in_program" in case resuming incomplete scraping
        for website, count in self.url_dict.items():
            i= count
            #update scraping function here....
            try:
                if(j>=k):
                    webbrowser.open(self.list_of_links[j])    #open the first url of each website to edit the scraping function accordingly
            except:
                print(f"Website didn't open! Open this link: {self.list_of_links[j]}")   
            if(j>=k):
                print(f"Website currently being scraped: {website}, Count: {count}")
                index = list(self.url_dict.keys()).index(website)
                print(f"The index of website {website} is {index}")
                checkinput = input("Press enter once you're done changing the scraping function. Press 0 to exit.")
                if(checkinput=="0"):
                    break
                reload(scrape_text)

            #use scraping function to scrape for each website
            
            while(i>0):
                
                if(k>j):
                    i-=1
                    j+=1
                    continue
                flag = scrape_text.scrape_text(j, self.list_of_links[j])
                if(flag):
                    link_exceptions.append(self.list_of_links[j])
                # print(flag)
                i-=1
                j+=1
                with open('position_in_program.txt', 'w') as f1:
                    f1.write(str(j))
        with open(os.path.join('exceptionlinks.txt'), 'a') as f:
            for exceptionlink in link_exceptions:
                f.write(exceptionlink)
                f.write('\n')

datascraper = DataScraperFromLinks()
datascraper.copy_links_from_excel_and_scrape_text(excel_filename= 'HINDI-HEMANTHA.xlsx')